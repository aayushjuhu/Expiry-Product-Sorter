from flask import Flask,render_template,request
import openpyxl
from openpyxl.styles import Border,Side,Font,Alignment
import webbrowser

app=Flask(__name__)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/send', methods=['POST'])
def send():
    if request.method == 'POST':
        pname = request.form.get('pname')
        bno = request.form.get('bno')
        edt = request.form.get('edt')
        qty = request.form.get('qty')
        pkg = request.form.get('pkg')
        mrp = request.form.get('mrp')
        pr = request.form.get('pr')
        dname = request.form.get('dname')
        print(f'{pname},{bno},{edt},{qty},{pkg},{mrp},{pr},{dname}')

        # Define file path
        fp = f'excel/{dname}.xlsx'
        curr_wbk = openpyxl.load_workbook(fp)
        curr_sheet = curr_wbk.active

        # Store merged cell ranges before inserting a row
        merged_ranges = list(curr_sheet.merged_cells.ranges)
        merged_cell_text = {}

        for merged_range in merged_ranges:
            start_cell = merged_range.start_cell
            merged_cell_text[str(merged_range)] = start_cell.value  # Store text from the first merged cell
            curr_sheet.unmerge_cells(str(merged_range))  # Unmerge before inserting

        insert_row = 11

        # Insert a new row at row 11
        curr_sheet.insert_rows(insert_row)

        
        # Update merged cell references
        new_merged_ranges = []
        for merged_range in merged_ranges:
            start_cell, end_cell = str(merged_range).split(":")
            
            # Extract start and end row numbers
            start_col = "".join(filter(str.isalpha, start_cell))
            start_row = int("".join(filter(str.isdigit, start_cell)))
            end_col = "".join(filter(str.isalpha, end_cell))
            end_row = int("".join(filter(str.isdigit, end_cell)))

            # Shift down merged ranges if they are at or below row 11
            if start_row >= insert_row:
                start_row += 1
            if end_row >= insert_row:
                end_row += 1

            # Construct the new merged range
            new_range = f"{start_col}{start_row}:{end_col}{end_row}"
            new_merged_ranges.append(new_range)
            curr_sheet.merge_cells(new_range)
            
            first_cell = curr_sheet[f"{start_col}{start_row}"]
            first_cell.value = merged_cell_text.get(str(merged_range), "")  # Restore stored text
            first_cell.alignment = Alignment(horizontal="center", vertical="center")  # Keep formatting


            r1="Signature & Stamp of Retailer"
            r2='Signature of Receipent 1'
            r3='Signature of Receipent 2'
            r4='Leakage/Breakage/Expired Drugs Not For Sale And Composition'
            r5=curr_sheet['A3']

            for row in curr_sheet.iter_rows():
                for cell in row:
                    if cell.value == r1:
                        # Apply formatting to the matching cell
                        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                        print(f"Formatted cell found at {cell.coordinate}")  # Debugging output
                    if cell.value == r2:
                        # Apply formatting to the matching cell
                        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                        print(f"Formatted cell found at {cell.coordinate}")  # Debugging output
                    if cell.value == r3:
                        # Apply formatting to the matching cell
                        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                        print(f"Formatted cell found at {cell.coordinate}")  # Debugging output
                    if cell.value == r4:
                        # Apply formatting to the matching cell
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        print(f"Formatted cell found at {cell.coordinate}")  # Debugging output
                    

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        font_style = Font(name="Calibri", size=12, bold=True)
        product_data = [pname, bno, "", edt, qty, pkg, mrp, "", "", "", "", "", "", "", "", "", "", pr]
        for col_num, value in enumerate(product_data, start=1):
            cell = curr_sheet.cell(row=insert_row, column=col_num, value=value)

            cell.border = thin_border
            cell.font = font_style

        r5.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        curr_wbk.save(fp)
        curr_wbk.close()
        return "<script>alert('Data saved successfully!'); window.location.href='/';</script>"
    else:
        return "Forbidden Access"



if __name__=='__main__':
    webbrowser.open_new_tab("http://127.0.0.1:5000/")
    app.run(debug=False, port=5000)
